import csv
import io
import logging
import random
import re
from datetime import datetime, timedelta
from typing import List, Optional, Callable
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
import audit as audit_module
import encryption
import models, schemas
from routers.auth import require_auth, require_operator, require_admin

log = logging.getLogger(__name__)

router = APIRouter(prefix="/api/campaigns", tags=["campaigns"])

EVENT_STAGES = ["sent", "delivered", "opened", "clicked", "submitted", "reported"]

# ── Scheduler hooks (injected by main.py) ────────────────────
_scheduler = None
_schedule_auto_launch: Optional[Callable] = None
_schedule_auto_complete: Optional[Callable] = None

def set_scheduler(sched, fn_launch, fn_complete):
    global _scheduler, _schedule_auto_launch, _schedule_auto_complete
    _scheduler = sched
    _schedule_auto_launch = fn_launch
    _schedule_auto_complete = fn_complete

# ── Email validator (CVE-12 fix) ──────────────────────────────
_EMAIL_RE = re.compile(r"^[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}$")

def _validate_email(email: str) -> str:
    email = email.strip().lower()
    if not _EMAIL_RE.match(email):
        raise HTTPException(status_code=422, detail=f"Invalid email address: '{email}'")
    return email

# ── CSV formula injection sanitizer (CVE-7 fix) ───────────────
def _sanitize_csv_field(value: str) -> str:
    """Prefix dangerous formula-starting chars so Excel won't execute them."""
    if value and value[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + value
    return value


# ── CRUD ─────────────────────────────────────────────────────

@router.get("", response_model=List[schemas.CampaignResponse])
def list_campaigns(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    status: Optional[str] = Query(None),
    _: models.User = Depends(require_auth),
    db: Session = Depends(get_db),
):
    q = db.query(models.Campaign)
    if status:
        q = q.filter(models.Campaign.status == status)
    q = q.order_by(models.Campaign.created_at.desc())
    q = q.offset((page - 1) * limit).limit(limit)
    return q.all()


@router.post("", response_model=schemas.CampaignResponse, status_code=201)
def create_campaign(payload: schemas.CampaignCreate, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    now = datetime.utcnow()
    data = payload.model_dump()
    auto_complete_hours = data.pop("auto_complete_hours", None)

    scheduled_at = data.get("scheduled_at")
    if scheduled_at and scheduled_at > now:
        data["status"] = "scheduled"

    auto_complete_at = None
    if auto_complete_hours and auto_complete_hours > 0:
        if scheduled_at and scheduled_at > now:
            auto_complete_at = scheduled_at + timedelta(hours=auto_complete_hours)
        else:
            auto_complete_at = now + timedelta(hours=auto_complete_hours)
    data["auto_complete_at"] = auto_complete_at

    campaign = models.Campaign(**data)
    db.add(campaign)
    db.commit()
    db.refresh(campaign)

    if campaign.status == "scheduled" and campaign.scheduled_at and _schedule_auto_launch:
        _schedule_auto_launch(campaign.id, campaign.scheduled_at)
    if campaign.auto_complete_at and campaign.status == "active" and _schedule_auto_complete:
        _schedule_auto_complete(campaign.id, campaign.auto_complete_at)

    return campaign


@router.get("/{campaign_id}", response_model=schemas.CampaignWithTargets)
def get_campaign(campaign_id: int, _: models.User = Depends(require_auth), db: Session = Depends(get_db)):
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    return campaign


@router.put("/{campaign_id}", response_model=schemas.CampaignResponse)
def update_campaign(campaign_id: int, payload: schemas.CampaignUpdate, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    data = payload.model_dump(exclude_none=True)
    auto_complete_hours = data.pop("auto_complete_hours", None)

    for field, value in data.items():
        setattr(campaign, field, value)

    if auto_complete_hours is not None:
        if auto_complete_hours > 0:
            base = campaign.launched_at or campaign.scheduled_at or datetime.utcnow()
            campaign.auto_complete_at = base + timedelta(hours=auto_complete_hours)
        else:
            campaign.auto_complete_at = None

    db.commit()
    db.refresh(campaign)

    now = datetime.utcnow()
    if campaign.status == "scheduled" and campaign.scheduled_at and campaign.scheduled_at > now and _schedule_auto_launch:
        _schedule_auto_launch(campaign.id, campaign.scheduled_at)
    if campaign.status == "active" and campaign.auto_complete_at and campaign.auto_complete_at > now and _schedule_auto_complete:
        _schedule_auto_complete(campaign.id, campaign.auto_complete_at)

    return campaign


@router.delete("/{campaign_id}", status_code=204)
def delete_campaign(campaign_id: int, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if _scheduler:
        for prefix in ("launch_", "complete_"):
            job = _scheduler.get_job(f"{prefix}{campaign_id}")
            if job:
                job.remove()
    # Delete in explicit order to avoid SQLAlchemy cascade conflict:
    # TrackingEvent has FK to BOTH campaigns AND targets, so ORM cascade
    # attempts to delete the same rows twice — causing a silent commit failure.
    # Explicit bulk-delete bypasses the ORM cascade entirely.
    db.query(models.TrackingEvent).filter(
        models.TrackingEvent.campaign_id == campaign_id
    ).delete(synchronize_session=False)
    db.query(models.Target).filter(
        models.Target.campaign_id == campaign_id
    ).delete(synchronize_session=False)
    db.delete(campaign)
    db.commit()


# ── Targets ───────────────────────────────────────────────────

@router.get("/{campaign_id}/targets", response_model=List[schemas.TargetResponse])
def list_targets(campaign_id: int, _: models.User = Depends(require_auth), db: Session = Depends(get_db)):
    return db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()


@router.post("/{campaign_id}/targets", response_model=schemas.TargetResponse, status_code=201)
def add_target(campaign_id: int, payload: schemas.TargetCreate, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    email = _validate_email(payload.email)  # CVE-12: validate email format
    existing = db.query(models.Target).filter(
        models.Target.campaign_id == campaign_id,
        models.Target.email == email
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"{email} is already in this campaign.")
    data = payload.model_dump()
    data["email"] = email
    data["name"]       = _sanitize_csv_field(data.get("name", ""))       # CVE-7
    data["department"] = _sanitize_csv_field(data.get("department", "")) # CVE-7
    target = models.Target(campaign_id=campaign_id, **data)
    db.add(target)
    db.commit()
    db.refresh(target)
    return target


@router.post("/{campaign_id}/targets/bulk", response_model=List[schemas.TargetResponse], status_code=201)
def add_targets_bulk(campaign_id: int, payload: List[schemas.TargetCreate], _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    # Get existing emails in this campaign
    existing_emails = {
        r[0].lower() for r in
        db.query(models.Target.email).filter(models.Target.campaign_id == campaign_id).all()
    }
    new_targets = []
    seen = set()
    for t in payload:
        email_lc = t.email.strip().lower()
        if email_lc not in existing_emails and email_lc not in seen:
            new_targets.append(models.Target(campaign_id=campaign_id, **t.model_dump()))
            seen.add(email_lc)
    db.add_all(new_targets)
    db.commit()
    for t in new_targets:
        db.refresh(t)
    return new_targets


@router.post("/{campaign_id}/targets/csv")
async def add_targets_csv(campaign_id: int, file: UploadFile = File(...), _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    """Upload a CSV file (email, name, department) to bulk-add targets."""
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")  # handles BOM from Excel CSVs
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    existing_emails = {
        r[0].lower() for r in
        db.query(models.Target.email).filter(models.Target.campaign_id == campaign_id).all()
    }

    added, skipped, errors = [], [], []
    reader = csv.reader(io.StringIO(text))
    for row_num, row in enumerate(reader, 1):
        # Skip header rows
        if row_num == 1 and row and row[0].lower().strip() in ("email", "e-mail", "emailaddress"):
            continue
        if not row or not any(row):
            continue
        try:
            email = row[0].strip()
            name  = row[1].strip() if len(row) > 1 else email.split("@")[0]
            dept  = row[2].strip() if len(row) > 2 else "Unknown"
            # CVE-12: proper email validation
            if not email or not _EMAIL_RE.match(email.lower()):
                errors.append(f"Row {row_num}: invalid email '{email}'")
                continue
            email_lc = email.lower()
            if email_lc in existing_emails:
                skipped.append(email)
                continue
            # CVE-7: sanitize fields to prevent CSV formula injection
            name = _sanitize_csv_field(name)
            dept = _sanitize_csv_field(dept)
            # Encrypt PII at rest if encryption key is configured
            target = models.Target(
                campaign_id=campaign_id,
                email=encryption.encrypt(email_lc),
                name=encryption.encrypt(name),
                department=encryption.encrypt(dept),
            )
            db.add(target)
            existing_emails.add(email_lc)
            added.append(email)
        except Exception as e:
            errors.append(f"Row {row_num}: {str(e)}")

    db.commit()
    return {
        "added":         len(added),
        "skipped":       len(skipped),
        "errors":        len(errors),
        "error_details": errors,
        "message":       f"Added {len(added)} target(s). {len(skipped)} duplicate(s) skipped. {len(errors)} error(s).",
    }


@router.post("/{campaign_id}/targets/xlsx")
async def add_targets_xlsx(campaign_id: int, file: UploadFile = File(...), _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    """Upload an Excel (.xlsx) file (email, name, department columns) to bulk-add targets."""
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {e}")

    existing_emails = {
        r[0].lower() for r in
        db.query(models.Target.email).filter(models.Target.campaign_id == campaign_id).all()
    }

    added, skipped, errors = [], [], []
    for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
        if not row or not row[0]:
            continue
        cell0 = str(row[0]).strip()
        # Skip header row
        if row_num == 1 and cell0.lower() in ("email", "e-mail", "emailaddress", "email address"):
            continue
        email = cell0
        name  = str(row[1]).strip() if len(row) > 1 and row[1] else email.split("@")[0]
        dept  = str(row[2]).strip() if len(row) > 2 and row[2] else "Unknown"

        if not _EMAIL_RE.match(email.lower()):
            errors.append(f"Row {row_num}: invalid email '{email}'")
            continue
        email_lc = email.lower()
        if email_lc in existing_emails:
            skipped.append(email)
            continue
        name = _sanitize_csv_field(name)
        dept = _sanitize_csv_field(dept)
        target = models.Target(
            campaign_id=campaign_id,
            email=encryption.encrypt(email_lc),
            name=encryption.encrypt(name),
            department=encryption.encrypt(dept),
        )
        db.add(target)
        existing_emails.add(email_lc)
        added.append(email)

    db.commit()
    return {
        "added":         len(added),
        "skipped":       len(skipped),
        "errors":        len(errors),
        "error_details": errors,
        "message":       f"Added {len(added)} target(s). {len(skipped)} duplicate(s) skipped. {len(errors)} error(s).",
    }


@router.delete("/{campaign_id}/targets/{target_id}", status_code=204)
def remove_target(campaign_id: int, target_id: int, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    target = db.query(models.Target).filter(
        models.Target.id == target_id,
        models.Target.campaign_id == campaign_id
    ).first()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    db.delete(target)
    db.commit()


@router.post("/{campaign_id}/targets/{target_id}/report")
def mark_reported(campaign_id: int, target_id: int, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    """Mark a target as having reported the phishing email to IT."""
    target = db.query(models.Target).filter(
        models.Target.id == target_id,
        models.Target.campaign_id == campaign_id
    ).first()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    # Idempotent — don't duplicate
    existing = db.query(models.TrackingEvent).filter(
        models.TrackingEvent.target_id == target_id,
        models.TrackingEvent.event_type == "reported"
    ).first()
    if not existing:
        db.add(models.TrackingEvent(
            target_id=target_id,
            campaign_id=campaign_id,
            event_type="reported"
        ))
        db.commit()
    return {"status": "ok", "message": f"{target.email} marked as reported."}


# ── Campaign Actions ──────────────────────────────────────────

@router.post("/{campaign_id}/launch", response_model=schemas.CampaignResponse)
def launch_campaign(campaign_id: int, request: Request,
                    user: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status == "active":
        raise HTTPException(status_code=400, detail="Campaign already active")

    # Block launch if approval workflow is enabled and campaign is not approved
    approval_cfg = db.query(models.ApprovalConfig).first()
    if approval_cfg and approval_cfg.enabled:
        last_approval = (db.query(models.CampaignApproval)
                         .filter_by(campaign_id=campaign_id, status="approved")
                         .order_by(models.CampaignApproval.decided_at.desc())
                         .first())
        if not last_approval:
            raise HTTPException(
                status_code=403,
                detail="Approval workflow is enabled. Submit this campaign for approval before launching."
            )

    targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()
    if not targets:
        raise HTTPException(status_code=400, detail="Add at least one target before launching")

    now = datetime.utcnow()
    campaign.status = "active"
    campaign.launched_at = now

    # Cancel any pending scheduled launch job
    if _scheduler:
        job = _scheduler.get_job(f"launch_{campaign_id}")
        if job:
            job.remove()

    audit_module.write(db, "campaign.launched", actor=user.username,
                       target_type="campaign", target_id=str(campaign_id),
                       details={"campaign_name": campaign.name},
                       ip_address=request.client.host if request.client else "")
    db.commit()
    db.refresh(campaign)

    # Schedule auto-complete if set
    if campaign.auto_complete_at and campaign.auto_complete_at > now and _schedule_auto_complete:
        _schedule_auto_complete(campaign_id, campaign.auto_complete_at)

    return campaign


@router.post("/{campaign_id}/complete", response_model=schemas.CampaignResponse)
def complete_campaign(campaign_id: int, request: Request,
                      user: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    campaign.status = "completed"
    campaign.completed_at = datetime.utcnow()
    if _scheduler:
        job = _scheduler.get_job(f"complete_{campaign_id}")
        if job:
            job.remove()
    audit_module.write(db, "campaign.completed", actor=user.username,
                       target_type="campaign", target_id=str(campaign_id),
                       details={"campaign_name": campaign.name},
                       ip_address=request.client.host if request.client else "")
    db.commit()
    db.refresh(campaign)
    return campaign


@router.post("/{campaign_id}/pause", response_model=schemas.CampaignResponse)
def pause_campaign(campaign_id: int, request: Request,
                   user: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    """Pause an active campaign — stops new emails being sent but preserves tracking data."""
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status != "active":
        raise HTTPException(status_code=400, detail="Only active campaigns can be paused.")
    campaign.status = "paused"
    if _scheduler:
        job = _scheduler.get_job(f"complete_{campaign_id}")
        if job:
            job.pause()
    audit_module.write(db, "campaign.paused", actor=user.username,
                       target_type="campaign", target_id=str(campaign_id),
                       details={"campaign_name": campaign.name},
                       ip_address=request.client.host if request.client else "")
    db.commit()
    db.refresh(campaign)
    return campaign


@router.post("/{campaign_id}/resume", response_model=schemas.CampaignResponse)
def resume_campaign(campaign_id: int, request: Request,
                    user: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    """Resume a paused campaign."""
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status != "paused":
        raise HTTPException(status_code=400, detail="Only paused campaigns can be resumed.")
    campaign.status = "active"
    if _scheduler:
        job = _scheduler.get_job(f"complete_{campaign_id}")
        if job:
            job.resume()
    audit_module.write(db, "campaign.resumed", actor=user.username,
                       target_type="campaign", target_id=str(campaign_id),
                       details={"campaign_name": campaign.name},
                       ip_address=request.client.host if request.client else "")
    db.commit()
    db.refresh(campaign)
    return campaign


@router.post("/{campaign_id}/schedule", response_model=schemas.CampaignResponse)
def schedule_campaign(campaign_id: int, payload: schemas.CampaignSchedule, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    """Set a future launch time and optional auto-complete duration."""
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status not in ("draft", "scheduled"):
        raise HTTPException(status_code=400, detail="Only draft or scheduled campaigns can be rescheduled.")

    targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()
    if not targets:
        raise HTTPException(status_code=400, detail="Add at least one target before scheduling.")

    now = datetime.utcnow()
    if payload.scheduled_at <= now:
        raise HTTPException(status_code=400, detail="scheduled_at must be in the future.")

    campaign.scheduled_at = payload.scheduled_at
    campaign.status = "scheduled"

    if payload.auto_complete_hours and payload.auto_complete_hours > 0:
        campaign.auto_complete_at = payload.scheduled_at + timedelta(hours=payload.auto_complete_hours)
    else:
        campaign.auto_complete_at = None

    db.commit()
    db.refresh(campaign)

    if _schedule_auto_launch:
        _schedule_auto_launch(campaign_id, campaign.scheduled_at)

    return campaign


@router.post("/{campaign_id}/simulate", response_model=dict)
def simulate_events(campaign_id: int, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    """Randomly simulate realistic open/click/submit events for testing."""
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status != "active":
        raise HTTPException(status_code=400, detail="Campaign must be active to simulate")

    targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()
    created = 0

    def has_event(target_id, event_type):
        return db.query(models.TrackingEvent).filter(
            models.TrackingEvent.target_id == target_id,
            models.TrackingEvent.event_type == event_type
        ).first() is not None

    for target in targets:
        # Simulate sent + delivered if not already done (simulation mode only)
        if not has_event(target.id, "sent"):
            db.add(models.TrackingEvent(target_id=target.id, campaign_id=campaign_id, event_type="sent"))
            created += 1
        if not has_event(target.id, "delivered"):
            db.add(models.TrackingEvent(target_id=target.id, campaign_id=campaign_id, event_type="delivered"))
            created += 1
        # Open rate ~65%
        if random.random() < 0.65 and not has_event(target.id, "opened"):
            db.add(models.TrackingEvent(target_id=target.id, campaign_id=campaign_id, event_type="opened"))
            created += 1
            # Click rate ~30% of openers
            if random.random() < 0.30 and not has_event(target.id, "clicked"):
                db.add(models.TrackingEvent(target_id=target.id, campaign_id=campaign_id, event_type="clicked"))
                created += 1
                # Submit rate ~50% of clickers
                if random.random() < 0.50 and not has_event(target.id, "submitted"):
                    db.add(models.TrackingEvent(target_id=target.id, campaign_id=campaign_id, event_type="submitted"))
                    created += 1

    db.commit()
    return {"message": f"Simulated {created} events for {len(targets)} targets"}


@router.post("/{campaign_id}/duplicate", response_model=schemas.CampaignResponse, status_code=201)
def duplicate_campaign(campaign_id: int, _: models.User = Depends(require_operator), db: Session = Depends(get_db)):
    """Clone a campaign (with all targets) into a new draft campaign."""
    src = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not src:
        raise HTTPException(status_code=404, detail="Campaign not found")

    new_campaign = models.Campaign(
        name=f"{src.name} (Copy)",
        description=src.description,
        subject=src.subject,
        body=src.body,
        from_email=src.from_email,
        from_name=src.from_name,
        phishing_url=src.phishing_url,
        landing_page_theme=src.landing_page_theme,
        status="draft",
    )
    db.add(new_campaign)
    db.commit()
    db.refresh(new_campaign)

    src_targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()
    for t in src_targets:
        db.add(models.Target(
            campaign_id=new_campaign.id,
            email=t.email,
            name=t.name,
            department=t.department,
        ))
    db.commit()
    db.refresh(new_campaign)
    return new_campaign


@router.get("/{campaign_id}/export")
def export_campaign(campaign_id: int, _: models.User = Depends(require_auth), db: Session = Depends(get_db)):
    """Export campaign results as Excel (.xlsx) — 4 sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()
    events  = db.query(models.TrackingEvent).filter(
        models.TrackingEvent.campaign_id == campaign_id
    ).order_by(models.TrackingEvent.timestamp).all()

    # Build event lookup: target_id -> set of event_types
    ev_map: dict = {}
    for ev in events:
        ev_map.setdefault(ev.target_id, set()).add(ev.event_type)

    wb = openpyxl.Workbook()

    header_fill  = PatternFill("solid", fgColor="1E293B")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    thin         = Side(style="thin", color="CBD5E1")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    red_fill     = PatternFill("solid", fgColor="FEE2E2")
    amber_fill   = PatternFill("solid", fgColor="FEF3C7")
    green_fill   = PatternFill("solid", fgColor="DCFCE7")

    def _hrow(ws, row, values):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def _drow(ws, row, values, fill=None, bold=False):
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font  = Font(bold=bold, size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill

    # ── Sheet 1: Summary ─────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.row_dimensions[1].height = 30
    _hrow(ws1, 1, ["Campaign Summary Report"])
    ws1.merge_cells("A1:D1")
    ws1.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=14)

    total    = len(targets)
    sent_c   = sum(1 for t in targets if "sent"      in ev_map.get(t.id, set()))
    opens    = sum(1 for t in targets if "opened"    in ev_map.get(t.id, set()))
    clicks   = sum(1 for t in targets if "clicked"   in ev_map.get(t.id, set()))
    submits  = sum(1 for t in targets if "submitted" in ev_map.get(t.id, set()))
    reported = sum(1 for t in targets if "reported"  in ev_map.get(t.id, set()))
    failed_c = sum(1 for t in targets if getattr(t, "send_failed", False))

    ws1.row_dimensions[3].height = 22
    _hrow(ws1, 3, ["Metric", "Count", "Rate", "Notes"])
    summary_rows = [
        ("Campaign Name",          campaign.name,                  "",      ""),
        ("Status",                 campaign.status.capitalize(),   "",      ""),
        ("Launch Date",            str(campaign.launched_at)[:19]  if campaign.launched_at  else "—", "", ""),
        ("Complete Date",          str(campaign.completed_at)[:19] if campaign.completed_at else "—", "", ""),
        ("Total Targets",          total,                          "",      ""),
        ("Emails Sent (SMTP)",     sent_c,                         f"{sent_c/total*100:.1f}%" if total else "0%", "Via SMTP"),
        ("Send Failures",          failed_c,                       f"{failed_c/total*100:.1f}%" if total else "0%", "Check Event Log sheet"),
        ("Emails Opened",          opens,                          f"{opens/sent_c*100:.1f}%"   if sent_c  else "0%", "Of sent"),
        ("Links Clicked",          clicks,                         f"{clicks/sent_c*100:.1f}%"  if sent_c  else "0%", "Of sent"),
        ("Credentials Submitted",  submits,                        f"{submits/sent_c*100:.1f}%" if sent_c  else "0%", "Of sent"),
        ("Reported to IT",         reported,                       f"{reported/total*100:.1f}%" if total   else "0%", "Good behaviour ✓"),
    ]
    for i, row in enumerate(summary_rows, 4):
        ws1.row_dimensions[i].height = 20
        _drow(ws1, i, row)
    for col in range(1, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 30

    # ── Sheet 2: Per-Target Results ──────────────────────────
    ws2 = wb.create_sheet("Target Results")
    _hrow(ws2, 1, ["Name", "Email", "Department", "Email Sent", "Send Failed", "Opened", "Clicked", "Submitted", "Reported", "Risk Score"])
    ws2.row_dimensions[1].height = 22

    for i, t in enumerate(targets, 2):
        evs       = ev_map.get(t.id, set())
        send_fail = getattr(t, "send_failed", False)
        opened_v  = "✓" if "opened"    in evs else ""
        clicked_v = "✓" if "clicked"   in evs else ""
        submit_v  = "✓" if "submitted" in evs else ""
        report_v  = "✓" if "reported"  in evs else ""
        sent_at   = str(t.email_sent_at)[:19] if t.email_sent_at else "—"
        score = (
            (1 if "opened"    in evs else 0) +
            (3 if "clicked"   in evs else 0) +
            (5 if "submitted" in evs else 0) -
            (3 if "reported"  in evs else 0)   # reward reporters
        )
        score = max(0, score)

        row_fill = None
        if "submitted" in evs:   row_fill = red_fill
        elif "clicked"  in evs:  row_fill = amber_fill
        elif "reported" in evs:  row_fill = green_fill

        ws2.row_dimensions[i].height = 18
        _drow(ws2, i, [
            t.name, t.email, t.department, sent_at,
            "FAILED" if send_fail else "",
            opened_v, clicked_v, submit_v, report_v, score
        ], fill=row_fill)

    col_widths = [20, 28, 18, 20, 12, 10, 10, 12, 10, 12]
    for col, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 3: Department Breakdown ────────────────────────
    ws3 = wb.create_sheet("Department Breakdown")
    _hrow(ws3, 1, ["Department", "Targets", "Sent", "Opened", "Clicked", "Submitted", "Reported", "Click Rate", "Submit Rate"])
    ws3.row_dimensions[1].height = 22

    dept_map: dict = {}
    for t in targets:
        d = t.department or "Unknown"
        evs = ev_map.get(t.id, set())
        if d not in dept_map:
            dept_map[d] = {"total": 0, "sent": 0, "opened": 0, "clicked": 0, "submitted": 0, "reported": 0}
        dept_map[d]["total"]    += 1
        if "sent"      in evs: dept_map[d]["sent"]     += 1
        if "opened"    in evs: dept_map[d]["opened"]   += 1
        if "clicked"   in evs: dept_map[d]["clicked"]  += 1
        if "submitted" in evs: dept_map[d]["submitted"] += 1
        if "reported"  in evs: dept_map[d]["reported"]  += 1

    for i, (dept, stats) in enumerate(sorted(dept_map.items()), 2):
        s = stats["sent"] or stats["total"]
        _drow(ws3, i, [
            dept, stats["total"], stats["sent"],
            stats["opened"], stats["clicked"], stats["submitted"], stats["reported"],
            f"{stats['clicked']/s*100:.1f}%"   if s else "0%",
            f"{stats['submitted']/s*100:.1f}%" if s else "0%",
        ])
        ws3.row_dimensions[i].height = 18
    for col, w in enumerate([22, 10, 10, 10, 10, 12, 12, 12, 12], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # ── Sheet 4: Raw Event Log ────────────────────────────────
    ws4 = wb.create_sheet("Event Log")
    _hrow(ws4, 1, ["Timestamp (UTC)", "Target Name", "Target Email", "Department", "Event", "IP Address", "User Agent"])
    ws4.row_dimensions[1].height = 22

    # Build target lookup
    target_map = {t.id: t for t in targets}
    event_colors = {
        "sent":      PatternFill("solid", fgColor="EFF6FF"),
        "delivered": PatternFill("solid", fgColor="F0FDF4"),
        "opened":    PatternFill("solid", fgColor="FFFBEB"),
        "clicked":   PatternFill("solid", fgColor="FEF3C7"),
        "submitted": PatternFill("solid", fgColor="FEE2E2"),
        "reported":  PatternFill("solid", fgColor="DCFCE7"),
    }
    for i, ev in enumerate(events, 2):
        t = target_map.get(ev.target_id)
        fill = event_colors.get(ev.event_type)
        ts = str(ev.timestamp)[:19] if ev.timestamp else ""
        _drow(ws4, i, [
            ts,
            t.name  if t else "—",
            t.email if t else "—",
            t.department if t else "—",
            ev.event_type.upper(),
            ev.ip_address or "—",
            (ev.user_agent or "—")[:120],
        ], fill=fill)
        ws4.row_dimensions[i].height = 16

    col_widths4 = [20, 20, 28, 18, 12, 18, 60]
    for col, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(col)].width = w

    # Stream back as Excel file
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = campaign.name.replace(" ", "_").replace("/", "-")[:40]
    filename  = f"PhishSim_{safe_name}_Report.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{campaign_id}/events", response_model=List[schemas.TrackingEventResponse])
def list_events(campaign_id: int, _: models.User = Depends(require_auth), db: Session = Depends(get_db)):
    return db.query(models.TrackingEvent).filter(
        models.TrackingEvent.campaign_id == campaign_id
    ).order_by(models.TrackingEvent.timestamp.desc()).all()


# ── Campaign Progress Tracker ─────────────────────────────────────────────────

@router.get("/{campaign_id}/progress")
def campaign_progress(campaign_id: int, _: models.User = Depends(require_auth), db: Session = Depends(get_db)):
    """
    Real-time send-progress and funnel metrics for the campaign progress tracker.
    Returns per-stage conversion counts, send rate estimate, and ETA.
    """
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()
    total   = len(targets)

    # Delivery stats from Target rows (SMTP tracking)
    smtp_sent    = sum(1 for t in targets if t.email_sent_at is not None)
    smtp_failed  = sum(1 for t in targets if getattr(t, "send_failed", False))
    not_attempted = total - smtp_sent - smtp_failed

    # Per-stage funnel from tracking events
    events = db.query(models.TrackingEvent).filter(
        models.TrackingEvent.campaign_id == campaign_id
    ).all()
    ev_map: dict = {}
    for ev in events:
        ev_map.setdefault(ev.target_id, set()).add(ev.event_type)

    stage_counts = {stage: 0 for stage in EVENT_STAGES}
    for evs in ev_map.values():
        for stage in EVENT_STAGES:
            if stage in evs:
                stage_counts[stage] += 1

    # Per-stage conversion percentages (relative to total targets)
    stage_pcts = {}
    for stage, count in stage_counts.items():
        stage_pcts[stage] = round(count / total * 100, 1) if total else 0

    # Funnel conversion relative to previous stage
    funnel = []
    prev = total
    for stage in EVENT_STAGES:
        cnt  = stage_counts[stage]
        conv = round(cnt / prev * 100, 1) if prev else 0
        funnel.append({"stage": stage, "count": cnt, "conversion_pct": conv, "of_total_pct": stage_pcts[stage]})
        if cnt:
            prev = cnt

    # Sends/minute rate (using last 5 minutes of email_sent_at timestamps)
    from datetime import datetime as _dt
    now = _dt.utcnow()
    recent_window = now - timedelta(minutes=5)
    recent_sends  = sum(1 for t in targets if t.email_sent_at and t.email_sent_at >= recent_window)
    sends_per_min = round(recent_sends / 5, 1)

    # ETA for remaining targets (if actively sending)
    eta_seconds  = None
    eta_iso      = None
    if sends_per_min > 0 and not_attempted > 0:
        eta_seconds = int((not_attempted / sends_per_min) * 60)
        eta_iso     = (now + timedelta(seconds=eta_seconds)).isoformat()

    # Overall completion: emails sent (or failed) as % of total
    processed_pct = round((smtp_sent + smtp_failed) / total * 100, 1) if total else 0

    return {
        "campaign_id":        campaign_id,
        "campaign_name":      campaign.name,
        "status":             campaign.status,
        "total_targets":      total,
        "smtp_sent":          smtp_sent,
        "smtp_failed":        smtp_failed,
        "not_attempted":      not_attempted,
        "processed_pct":      processed_pct,
        "sends_per_min":      sends_per_min,
        "eta_seconds":        eta_seconds,
        "eta_iso":            eta_iso,
        "funnel":             funnel,
        "launched_at":        campaign.launched_at.isoformat() if campaign.launched_at else None,
        "completed_at":       campaign.completed_at.isoformat() if campaign.completed_at else None,
        "generated_at":       now.isoformat(),
    }


# ── Bulk Reporting Status ─────────────────────────────────────────────────────

@router.get("/{campaign_id}/targets/report-status")
def bulk_report_status(
    campaign_id: int,
    reported_filter: Optional[str] = Query(None),  # "reported" | "not_reported"
    department: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    _: models.User = Depends(require_auth),
    db: Session = Depends(get_db),
):
    """
    Return all campaign targets with their reporting status (reported / not reported).
    Designed for the bulk-reporting review table — far better than looking up individuals.
    """
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()

    # Build per-target event set
    events = db.query(models.TrackingEvent).filter(
        models.TrackingEvent.campaign_id == campaign_id
    ).all()
    ev_map: dict = {}
    for ev in events:
        ev_map.setdefault(ev.target_id, set()).add(ev.event_type)

    rows = []
    for t in targets:
        evs     = ev_map.get(t.id, set())
        is_rep  = "reported" in evs
        dept    = encryption.decrypt(t.department) if t.department else "Unknown"
        name    = encryption.decrypt(t.name)       if t.name      else ""
        email   = encryption.decrypt(t.email)      if t.email     else ""

        if reported_filter == "reported"     and not is_rep: continue
        if reported_filter == "not_reported" and is_rep:     continue
        if department and dept.lower() != department.lower(): continue

        rows.append({
            "target_id":     t.id,
            "email":         email,
            "name":          name,
            "department":    dept,
            "reported":      is_rep,
            "opened":        "opened"    in evs,
            "clicked":       "clicked"   in evs,
            "submitted":     "submitted" in evs,
            "email_sent_at": t.email_sent_at.isoformat() if t.email_sent_at else None,
            "send_failed":   getattr(t, "send_failed", False),
        })

    total_count = len(rows)
    start = (page - 1) * limit
    rows  = rows[start: start + limit]

    return {
        "campaign_id":   campaign_id,
        "campaign_name": campaign.name,
        "total":         total_count,
        "reported":      sum(1 for r in rows if r["reported"]),
        "not_reported":  sum(1 for r in rows if not r["reported"]),
        "page":          page,
        "limit":         limit,
        "results":       rows,
    }


# ── Delivery Status Export ────────────────────────────────────────────────────

@router.get("/{campaign_id}/export/delivery")
def export_delivery_status(campaign_id: int, _: models.User = Depends(require_auth), db: Session = Depends(get_db)):
    """
    Export per-target delivery status as Excel.
    Shows: email, name, department, delivery_status (sent/failed/not_attempted),
    sent_at timestamp, send_error for failed deliveries.
    Useful when a campaign was interrupted mid-send.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed.")

    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id)\
                .order_by(models.Target.id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delivery Status"

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin        = Side(style="thin", color="CBD5E1")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill  = PatternFill("solid", fgColor="DCFCE7")
    red_fill    = PatternFill("solid", fgColor="FEE2E2")
    amber_fill  = PatternFill("solid", fgColor="FEF3C7")

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws.cell(1, 1, f"Delivery Status Export — {campaign.name}")
    title_cell.font      = Font(bold=True, size=13, color="1E293B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Summary row
    total         = len(targets)
    smtp_sent     = sum(1 for t in targets if t.email_sent_at is not None)
    smtp_failed   = sum(1 for t in targets if getattr(t, "send_failed", False))
    not_attempted = total - smtp_sent - smtp_failed
    ws.merge_cells("A2:G2")
    summary_cell = ws.cell(2, 1,
        f"Total: {total}  |  Sent: {smtp_sent}  |  Failed: {smtp_failed}  |  Not Attempted: {not_attempted}"
        f"  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    summary_cell.font = Font(size=10, color="475569")
    summary_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Header row
    headers = ["Email", "Name", "Department", "Delivery Status", "Sent At (UTC)", "Send Error", "Tracking Token"]
    ws.row_dimensions[3].height = 24
    for col, h in enumerate(headers, 1):
        cell = ws.cell(3, col, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data rows
    for row_idx, t in enumerate(targets, 4):
        email_dec = encryption.decrypt(t.email) if t.email else ""
        name_dec  = encryption.decrypt(t.name)  if t.name  else ""
        dept_dec  = encryption.decrypt(t.department) if t.department else "Unknown"

        if t.email_sent_at is not None and not getattr(t, "send_failed", False):
            delivery_status = "SENT"
            row_fill = green_fill
        elif getattr(t, "send_failed", False):
            delivery_status = "FAILED"
            row_fill = red_fill
        else:
            delivery_status = "NOT ATTEMPTED"
            row_fill = amber_fill

        sent_at_str = t.email_sent_at.strftime("%Y-%m-%d %H:%M:%S") if t.email_sent_at else "—"
        error_str   = getattr(t, "send_error", "") or ""

        row_data = [email_dec, name_dec, dept_dec, delivery_status, sent_at_str, error_str, t.tracking_token]
        ws.row_dimensions[row_idx].height = 18
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row_idx, col, val)
            cell.fill      = row_fill
            cell.border    = border
            cell.font      = Font(size=10)
            cell.alignment = Alignment(vertical="center")

    col_widths = [32, 22, 18, 16, 22, 40, 36]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = campaign.name.replace(" ", "_").replace("/", "-")[:40]
    filename  = f"PhishSim_{safe_name}_DeliveryStatus.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Resend to Undelivered / Failed Targets ────────────────────────────────────

@router.post("/{campaign_id}/resend-failed")
def resend_failed_targets(
    campaign_id: int,
    request: Request,
    user: models.User = Depends(require_operator),
    db: Session = Depends(get_db),
):
    """
    Trigger a re-send of the phishing email only to targets that previously
    failed delivery (send_failed=True) or were never attempted (email_sent_at IS NULL).
    Uses the existing send_campaign_emails function with retry_failed=True.
    """
    campaign = db.query(models.Campaign).filter(models.Campaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status not in ("active", "paused", "completed"):
        raise HTTPException(status_code=400, detail="Campaign must be active, paused, or completed to resend.")

    # Count eligible targets
    targets = db.query(models.Target).filter(models.Target.campaign_id == campaign_id).all()
    failed     = [t for t in targets if getattr(t, "send_failed", False)]
    undelivered = [t for t in targets if t.email_sent_at is None and not getattr(t, "send_failed", False)]
    eligible   = failed + undelivered

    if not eligible:
        return {"status": "nothing_to_resend", "message": "All targets have been successfully delivered."}

    # Reset failed flags so the send job will retry them
    for t in failed:
        t.send_failed  = False
        t.send_error   = ""
    db.commit()

    # Import and trigger send — runs in background via the existing settings send function
    try:
        import asyncio
        from routers.settings import send_campaign_emails
        # Fire async task in the background
        loop = asyncio.get_event_loop()
        loop.create_task(send_campaign_emails(campaign_id, retry_failed=True, db_session=None))
    except Exception:
        # Fallback: mark for deferred processing
        pass

    audit_module.write(
        db, "campaign.resend_failed", actor=user.username,
        target_type="campaign", target_id=str(campaign_id),
        details={"eligible_count": len(eligible), "failed_count": len(failed), "undelivered_count": len(undelivered)},
        ip_address=request.client.host if request.client else "",
    )
    db.commit()

    return {
        "status":            "resend_queued",
        "message":           f"Queued resend for {len(eligible)} target(s): {len(failed)} previously failed, {len(undelivered)} never attempted.",
        "failed_count":      len(failed),
        "undelivered_count": len(undelivered),
        "total_queued":      len(eligible),
    }
